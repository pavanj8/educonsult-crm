"""Analytics export endpoint for student lists (E44; Journey J37)."""

import csv
import io
from datetime import datetime
from io import StringIO
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import Session

from app.db.branch_scope import BranchScopeError, apply_branch_scope
from app.db.database import get_db
from app.db.tenant_scope import TenantScopeError, apply_tenant_scope
from app.models.branch import Branch
from app.models.user import User
from app.rbac.dependencies import require_permission
from app.rbac.permissions import Permission
from app.rbac.roles import Role
from app.rbac.user import AuthenticatedUser

router = APIRouter()
_DB_UNAVAILABLE_DETAIL = "Analytics service is temporarily unavailable"


@router.get("/students")
def export_student_list(
    current_user: Annotated[
        AuthenticatedUser,
        Depends(
            require_permission(Permission.REPORT_EXPORT),
        ),
    ],
    format: Annotated[
        str,
        Query(
            description="Export format: 'csv' or 'xlsx'",
            pattern="^(csv|xlsx)$",
        ),
    ] = "csv",
    start_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter students created on or after this date/time "
                "(ISO 8601 format). Defaults to beginning of available data."
            )
        ),
    ] = None,
    end_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter students created before or on this date/time "
                "(ISO 8601 format). Defaults to current time."
            )
        ),
    ] = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export student list to CSV or Excel (E44; Journey J37).

    Returns a downloadable file containing all students scoped to the
    caller's tenant/branch, with optional date range filtering on
    student creation date.

    **Permission**: ``REPORT_EXPORT`` (branch manager and above)

    **Scoping**:
    - Branch managers see only students from their assigned branch
    - Consultancy owners see all branches in their tenant
    - Super admins see all tenants (platform-wide view)

    **Date filtering**:
    - Both ``start_date`` and ``end_date`` are optional
    - When provided, filters students by ``created_at`` timestamp
    - ``start_date`` is inclusive (>=), ``end_date`` is inclusive (<=)

    **Export columns**:
    - Student ID
    - Email
    - Name
    - Phone
    - Date of Birth
    - Branch Name
    - Branch City
    - Target Country (ID or empty)
    - Target University (ID or empty)
    - Target Program (ID or empty)
    - Created At (ISO 8601)
    - Is Active

    **Response**:
    - CSV: ``text/csv`` with header row
    - Excel (``xlsx``): ``application/vnd.openxmlformats-officedocument.spreadsheetml.sheet``
    - filename includes ``students-`` and timestamp
    """
    try:
        import csv
        from io import StringIO

        # Build the base query with tenant scoping - filter for STUDENT role only
        statement = apply_tenant_scope(select(User), User, current_user)

        # Apply branch scoping (Branch Manager sees only their branch, Owner sees all)
        statement = apply_branch_scope(statement, User, current_user)

        # Filter for students only (role = STUDENT)
        statement = statement.where(User.role == Role.STUDENT)

        # Apply date range filters if provided
        if start_date is not None:
            statement = statement.where(User.created_at >= start_date)
        if end_date is not None:
            statement = statement.where(User.created_at <= end_date)

        # Join with Branch to get branch name and city
        statement = statement.outerjoin(Branch, User.branch_id == Branch.id).add_columns(
            User.id,
            User.email,
            User.name,
            User.phone,
            User.date_of_birth,
            User.target_country_id,
            User.target_university_id,
            User.target_program_id,
            User.created_at,
            User.is_active,
            Branch.name.label("branch_name"),
            Branch.city.label("branch_city"),
        )

        # Order by creation date (newest first)
        statement = statement.order_by(User.created_at.desc())

        result = db.execute(statement).all()

        if format == "csv":
            # Generate CSV content
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(
                [
                    "Student ID",
                    "Email",
                    "Name",
                    "Phone",
                    "Date of Birth",
                    "Branch Name",
                    "Branch City",
                    "Target Country ID",
                    "Target University ID",
                    "Target Program ID",
                    "Created At",
                    "Is Active",
                ]
            )

            # Write data rows
            for row in result:
                writer.writerow(
                    [
                        row.id,
                        row.email,
                        row.name or "",
                        row.phone or "",
                        str(row.date_of_birth) if row.date_of_birth else "",
                        row.branch_name or "",
                        row.branch_city or "",
                        row.target_country_id or "",
                        row.target_university_id or "",
                        row.target_program_id or "",
                        row.created_at.isoformat() if row.created_at else "",
                        "Yes" if row.is_active else "No",
                    ]
                )

            # Reset pointer to beginning
            output.seek(0)

            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"students-{timestamp}.csv"

            return StreamingResponse(
                output,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )

        else:  # format == "xlsx"
            import io

            try:
                from openpyxl import Workbook
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail="Excel export requires openpyxl library to be installed",
                ) from None

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"

            # Write header row
            headers = [
                "Student ID",
                "Email",
                "Name",
                "Phone",
                "Date of Birth",
                "Branch Name",
                "Branch City",
                "Target Country ID",
                "Target University ID",
                "Target Program ID",
                "Created At",
                "Is Active",
            ]
            ws.append(headers)

            # Write data rows
            for row in result:
                ws.append(
                    [
                        row.id,
                        row.email,
                        row.name or "",
                        row.phone or "",
                        str(row.date_of_birth) if row.date_of_birth else "",
                        row.branch_name or "",
                        row.branch_city or "",
                        row.target_country_id or "",
                        row.target_university_id or "",
                        row.target_program_id or "",
                        row.created_at.isoformat() if row.created_at else "",
                        "Yes" if row.is_active else "No",
                    ]
                )

            # Save to memory
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"students-{timestamp}.xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )

    except (TenantScopeError, BranchScopeError):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions",
        ) from None
    except OperationalError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=_DB_UNAVAILABLE_DETAIL,
        ) from None


@router.get("/funnel")
def export_conversion_funnel(
    current_user: Annotated[
        AuthenticatedUser,
        Depends(
            require_permission(Permission.REPORT_EXPORT),
        ),
    ],
    format: Annotated[
        str,
        Query(
            description="Export format: 'csv' or 'xlsx'",
            pattern="^(csv|xlsx)$",
        ),
    ] = "csv",
    start_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter applications created on or after this date/time "
                "(ISO 8601 format). Defaults to beginning of available data."
            )
        ),
    ] = None,
    end_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter applications created before or on this date/time "
                "(ISO 8601 format). Defaults to current time."
            )
        ),
    ] = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export conversion funnel data to CSV or Excel (E44; Journey J37).

    Returns a downloadable file containing the conversion funnel metrics
    showing the count of applications at each pipeline stage, scoped to the
    caller's tenant/branch, with optional date range filtering on application
    creation date.

    **Permission**: ``REPORT_EXPORT`` (branch manager and above)

    **Scoping**:
    - Branch managers see only applications from their assigned branch
    - Consultancy owners see all branches in their tenant
    - Super admins see all tenants (platform-wide view)

    **Export columns**:
    - Stage
    - Count
    - Percentage
    """
    try:
        from app.models.application import Application
        from app.pipeline.stages import Stage

        # Build the base query with tenant and branch scoping
        statement = apply_tenant_scope(select(Application), Application, current_user)
        statement = apply_branch_scope(statement, Application, current_user)

        # Apply date range filters if provided
        if start_date is not None:
            statement = statement.where(Application.created_at >= start_date)
        if end_date is not None:
            statement = statement.where(Application.created_at <= end_date)

        # Get all applications
        result = db.execute(statement).scalars().all()

        # Count applications per stage
        stage_counts = {stage.value: 0 for stage in Stage}
        for app in result:
            stage_counts[app.stage] = stage_counts.get(app.stage, 0) + 1

        total = len(result)
        funnel_data = []
        for stage, count in stage_counts.items():
            percentage = (count / total * 100) if total > 0 else 0
            funnel_data.append({
                "stage": stage,
                "count": count,
                "percentage": f"{percentage:.2f}%",
            })

        if format == "csv":
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(["Stage", "Count", "Percentage"])

            # Write data rows
            for row in funnel_data:
                writer.writerow([row["stage"], row["count"], row["percentage"]])

            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"funnel-{timestamp}.csv"

            return StreamingResponse(
                output,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )
        else:  # format == "xlsx"
            try:
                from openpyxl import Workbook
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail="Excel export requires openpyxl library to be installed",
                ) from None

            wb = Workbook()
            ws = wb.active
            ws.title = "Conversion Funnel"

            # Write header row
            ws.append(["Stage", "Count", "Percentage"])

            # Write data rows
            for row in funnel_data:
                ws.append([row["stage"], row["count"], row["percentage"]])

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"funnel-{timestamp}.xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )

    except (TenantScopeError, BranchScopeError):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions",
        ) from None
    except OperationalError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=_DB_UNAVAILABLE_DETAIL,
        ) from None


@router.get("/registrations")
def export_registrations_over_time(
    current_user: Annotated[
        AuthenticatedUser,
        Depends(
            require_permission(Permission.REPORT_EXPORT),
        ),
    ],
    format: Annotated[
        str,
        Query(
            description="Export format: 'csv' or 'xlsx'",
            pattern="^(csv|xlsx)$",
        ),
    ] = "csv",
    start_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter students registered on or after this date/time "
                "(ISO 8601 format). Defaults to beginning of available data."
            )
        ),
    ] = None,
    end_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter students registered before or on this date/time "
                "(ISO 8601 format). Defaults to current time."
            )
        ),
    ] = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export registrations-over-time data to CSV or Excel (E44; Journey J37).

    Returns a downloadable file containing daily registration counts,
    scoped to the caller's tenant/branch, with optional date range filtering.

    **Permission**: ``REPORT_EXPORT`` (branch manager and above)

    **Export columns**:
    - Date
    - Count
    """
    try:
        from sqlalchemy import func

        # Build the base query with tenant and branch scoping
        statement = apply_tenant_scope(select(User), User, current_user)
        statement = apply_branch_scope(statement, User, current_user)

        # Filter for students only
        statement = statement.where(User.role == Role.STUDENT)

        # Apply date range filters if provided
        if start_date is not None:
            statement = statement.where(User.created_at >= start_date)
        if end_date is not None:
            statement = statement.where(User.created_at <= end_date)

        # Group by date and count
        statement = statement.add_columns(
            func.date(User.created_at).label("registration_date"),
            func.count(User.id).label("count")
        ).group_by(func.date(User.created_at)).order_by(func.date(User.created_at))

        result = db.execute(statement).all()

        if format == "csv":
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow(["Date", "Count"])

            # Write data rows
            for row in result:
                writer.writerow([str(row.registration_date), row.count])

            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"registrations-{timestamp}.csv"

            return StreamingResponse(
                output,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )
        else:  # format == "xlsx"
            try:
                from openpyxl import Workbook
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail="Excel export requires openpyxl library to be installed",
                ) from None

            wb = Workbook()
            ws = wb.active
            ws.title = "Registrations"

            # Write header row
            ws.append(["Date", "Count"])

            # Write data rows
            for row in result:
                ws.append([str(row.registration_date), row.count])

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"registrations-{timestamp}.xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )

    except (TenantScopeError, BranchScopeError):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions",
        ) from None
    except OperationalError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=_DB_UNAVAILABLE_DETAIL,
        ) from None


@router.get("/branch-comparison")
def export_branch_comparison(
    current_user: Annotated[
        AuthenticatedUser,
        Depends(
            require_permission(Permission.REPORT_EXPORT),
        ),
    ],
    format: Annotated[
        str,
        Query(
            description="Export format: 'csv' or 'xlsx'",
            pattern="^(csv|xlsx)$",
        ),
    ] = "csv",
    start_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter applications created on or after this date/time "
                "(ISO 8601 format). Defaults to beginning of available data."
            )
        ),
    ] = None,
    end_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter applications created before or on this date/time "
                "(ISO 8601 format). Defaults to current time."
            )
        ),
    ] = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export branch comparison data to CSV or Excel (E44; Journey J37).

    Returns a downloadable file containing branch-level metrics,
    scoped to the caller's tenant, with optional date range filtering.

    **Permission**: ``REPORT_EXPORT`` (consultancy owner and above)

    **Export columns**:
    - Branch Name
    - Branch City
    - Total Students
    - Active Applications
    - Enrolled Count
    - Rejected Count
    """
    try:
        from app.models.application import Application

        # Only owners and super admins can access branch comparison
        if current_user.role not in (Role.CONSULTANCY_OWNER, Role.SUPER_ADMIN):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Branch comparison export is only available to consultancy owners and super admins",
            )

        # Build base query for applications with tenant scoping
        app_statement = apply_tenant_scope(select(Application), Application, current_user)

        # Apply date range filters if provided
        if start_date is not None:
            app_statement = app_statement.where(Application.created_at >= start_date)
        if end_date is not None:
            app_statement = app_statement.where(Application.created_at <= end_date)

        # Get all applications
        applications = db.execute(app_statement).scalars().all()

        # Group by branch and calculate metrics
        from collections import defaultdict
        branch_metrics = defaultdict(lambda: {
            "total_students": set(),
            "active_applications": 0,
            "enrolled": 0,
            "rejected": 0,
        })

        for app in applications:
            branch_id = app.branch_id
            branch_metrics[branch_id]["total_students"].add(app.student_id)
            branch_metrics[branch_id]["active_applications"] += 1
            if app.stage == "enrolled":
                branch_metrics[branch_id]["enrolled"] += 1
            elif app.stage == "rejected":
                branch_metrics[branch_id]["rejected"] += 1

        # Get branch details
        branch_statement = apply_tenant_scope(select(Branch), Branch, current_user)
        branches = db.execute(branch_statement).scalars().all()
        branch_dict = {b.id: b for b in branches}

        # Build export data
        export_data = []
        for branch_id, metrics in branch_metrics.items():
            branch = branch_dict.get(branch_id)
            if branch:
                export_data.append({
                    "branch_name": branch.name,
                    "branch_city": branch.city,
                    "total_students": len(metrics["total_students"]),
                    "active_applications": metrics["active_applications"],
                    "enrolled": metrics["enrolled"],
                    "rejected": metrics["rejected"],
                })

        if format == "csv":
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                "Branch Name",
                "Branch City",
                "Total Students",
                "Active Applications",
                "Enrolled Count",
                "Rejected Count",
            ])

            # Write data rows
            for row in export_data:
                writer.writerow([
                    row["branch_name"],
                    row["branch_city"],
                    row["total_students"],
                    row["active_applications"],
                    row["enrolled"],
                    row["rejected"],
                ])

            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"branch-comparison-{timestamp}.csv"

            return StreamingResponse(
                output,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )
        else:  # format == "xlsx"
            try:
                from openpyxl import Workbook
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail="Excel export requires openpyxl library to be installed",
                ) from None

            wb = Workbook()
            ws = wb.active
            ws.title = "Branch Comparison"

            # Write header row
            ws.append([
                "Branch Name",
                "Branch City",
                "Total Students",
                "Active Applications",
                "Enrolled Count",
                "Rejected Count",
            ])

            # Write data rows
            for row in export_data:
                ws.append([
                    row["branch_name"],
                    row["branch_city"],
                    row["total_students"],
                    row["active_applications"],
                    row["enrolled"],
                    row["rejected"],
                ])

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"branch-comparison-{timestamp}.xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )

    except (TenantScopeError, BranchScopeError):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions",
        ) from None
    except OperationalError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=_DB_UNAVAILABLE_DETAIL,
        ) from None


@router.get("/platform-stats")
def export_platform_stats(
    current_user: Annotated[
        AuthenticatedUser,
        Depends(
            require_permission(Permission.REPORT_EXPORT),
        ),
    ],
    format: Annotated[
        str,
        Query(
            description="Export format: 'csv' or 'xlsx'",
            pattern="^(csv|xlsx)$",
        ),
    ] = "csv",
    start_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter users created on or after this date/time "
                "(ISO 8601 format). Defaults to beginning of available data."
            )
        ),
    ] = None,
    end_date: Annotated[
        datetime | None,
        Query(
            description=(
                "Filter users created before or on this date/time "
                "(ISO 8601 format). Defaults to current time."
            )
        ),
    ] = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Export platform-wide stats to CSV or Excel (E44; Journey J37).

    Returns a downloadable file containing platform-wide tenant statistics.
    Only available to super admins.

    **Permission**: ``REPORT_EXPORT`` (super admin only)

    **Export columns**:
    - Tenant Name
    - Plan
    - Total Branches
    - Total Students
    - Total Staff
    - Active Applications
    """
    try:
        from app.models.application import Application
        from app.models.tenant import Tenant

        # Only super admins can access platform stats
        if current_user.role != Role.SUPER_ADMIN:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Platform stats export is only available to super admins",
            )

        # Get all tenants
        tenant_statement = select(Tenant)
        tenants = db.execute(tenant_statement).scalars().all()

        # Build export data
        export_data = []
        for tenant in tenants:
            # Count branches for this tenant
            branch_count = db.execute(
                select(Branch).where(Branch.tenant_id == tenant.id)
            ).scalars().count()

            # Count students for this tenant
            student_count = db.execute(
                select(User).where(
                    User.tenant_id == tenant.id,
                    User.role == Role.STUDENT
                )
            ).scalars().count()

            # Count staff for this tenant
            staff_count = db.execute(
                select(User).where(
                    User.tenant_id == tenant.id,
                    User.role != Role.STUDENT,
                    User.role != Role.SUPER_ADMIN
                )
            ).scalars().count()

            # Count active applications for this tenant
            app_statement = select(Application).join(User, Application.student_id == User.id).where(
                User.tenant_id == tenant.id
            )

            # Apply date filters if provided
            if start_date is not None:
                app_statement = app_statement.where(Application.created_at >= start_date)
            if end_date is not None:
                app_statement = app_statement.where(Application.created_at <= end_date)

            active_applications = db.execute(app_statement).scalars().count()

            export_data.append({
                "tenant_name": tenant.name,
                "plan": tenant.plan or "N/A",
                "total_branches": branch_count,
                "total_students": student_count,
                "total_staff": staff_count,
                "active_applications": active_applications,
            })

        if format == "csv":
            output = StringIO()
            writer = csv.writer(output)

            # Write header
            writer.writerow([
                "Tenant Name",
                "Plan",
                "Total Branches",
                "Total Students",
                "Total Staff",
                "Active Applications",
            ])

            # Write data rows
            for row in export_data:
                writer.writerow([
                    row["tenant_name"],
                    row["plan"],
                    row["total_branches"],
                    row["total_students"],
                    row["total_staff"],
                    row["active_applications"],
                ])

            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"platform-stats-{timestamp}.csv"

            return StreamingResponse(
                output,
                media_type="text/csv",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )
        else:  # format == "xlsx"
            try:
                from openpyxl import Workbook
            except ImportError:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail="Excel export requires openpyxl library to be installed",
                ) from None

            wb = Workbook()
            ws = wb.active
            ws.title = "Platform Stats"

            # Write header row
            ws.append([
                "Tenant Name",
                "Plan",
                "Total Branches",
                "Total Students",
                "Total Staff",
                "Active Applications",
            ])

            # Write data rows
            for row in export_data:
                ws.append([
                    row["tenant_name"],
                    row["plan"],
                    row["total_branches"],
                    row["total_students"],
                    row["total_staff"],
                    row["active_applications"],
                ])

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"platform-stats-{timestamp}.xlsx"

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                },
            )

    except (TenantScopeError, BranchScopeError):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Insufficient permissions",
        ) from None
    except OperationalError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=_DB_UNAVAILABLE_DETAIL,
        ) from None
